# -*- coding: utf-8 -*-
"""人生选择清单 Excel 导出（数据来源：src/choices.js，经 choices-export.json）"""
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

data = json.load(open('choices-export.json', encoding='utf-8'))

BLACK = '000000'
GREY_BG = 'F5F5F5'
BORDER_GREY = 'D0D0D0'
BLUE = '0066CC'
BLUE_BG = 'E6F0FA'

thin = Font(size=11)
header_font = Font(size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')

CAT_DESC = {
    'A': '对而艰难的事 —— 提升难度同时提升收益',
    'B': '踏实稳定的事 —— 降低难度或稳定提升收益',
    'C': '出奇制胜的事 —— 改变规则',
    'D': '可遇不可求的事 —— 低概率（每格 8%）、高收益、无负面',
}

wb = Workbook()

# ─────────── Sheet 1: 封面 ───────────
ws = wb.active
ws.title = '封面'
ws.sheet_view.showGridLines = False
for col, w in zip('ABCDEFG', [3, 22, 30, 26, 22, 20, 3]):
    ws.column_dimensions[col].width = w

ws.merge_cells('B2:F2')
ws['B2'] = '人生一杯 · 人生选择配置清单'
ws['B2'].font = Font(size=20, bold=True, color=BLACK)
ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 38

ws.merge_cells('B3:F3')
ws['B3'] = '升段/升阶时三选一 · 数据来源：游戏配置 src/choices.js'
ws['B3'].font = Font(size=12, color='666666')
ws['B3'].alignment = Alignment(horizontal='center')

ws['B5'] = '内容概览（2026-08-20 按批注调整后）'
ws['B5'].font = Font(size=14, bold=True)
metrics = [
    ('A · 对而艰难的事', '16 个（删除 A6/A8/A9/A20；A3、A19 调整为每杯 +1 分）'),
    ('B · 踏实稳定的事', '18 个（删除 B7/B12）'),
    ('C · 出奇制胜的事', '14 个（删除 C3/C4/C12/C14/C19/C20；C6 调整为每杯 +1 分；C10 修复完成区显示 bug）'),
    ('D · 可遇不可求的事', '17 个（删除 D13/D14/D20/D21-23；D2 改 +40；D3 改 2 次；D8 上限 +50%；D11 改 +20）'),
    ('段位之力（TIER_FX）', '34 条（葡萄之力删除「前辈提点」；段位卡已下架，机制保留）'),
]
r = 6
for k, v in metrics:
    ws.cell(row=r, column=2, value=k).font = thin
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=v).font = Font(size=11, color=BLUE)
    r += 1

r += 1
ws.cell(row=r, column=2, value='工作表索引').font = Font(size=14, bold=True)
r += 1
for name, desc in [
    ('A 对而艰难的事', CAT_DESC['A']),
    ('B 踏实稳定的事', CAT_DESC['B']),
    ('C 出奇制胜的事', CAT_DESC['C']),
    ('D 可遇不可求的事', CAT_DESC['D']),
    ('段位之力', '各段位专属增益（段位卡已下架，效果池保留备用）'),
]:
    ws.cell(row=r, column=2, value=name).font = Font(size=11, bold=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=desc).font = thin
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws.cell(row=r, column=2, value='使用说明：每张表最后一列「调整意见」留空，可直接填写修改意见（改名/改数值/换专属段位/删除等），反馈后同步进游戏配置。').font = Font(size=10, color='888888')

# ─────────── Sheet 2-5: A/B/C/D 类选项 ───────────
HEADERS = ['编号', '选项名称', '效果说明', '生效阶段', '感悟文案', '调整意见']
WIDTHS = [8, 18, 38, 20, 40, 30]

def fill_cat_sheet(cat):
    title = cat + ' ' + CAT_DESC[cat].split(' —— ')[0]
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    for i, w in enumerate(WIDTHS):
        ws.column_dimensions[get_column_letter(i + 2)].width = w

    ws.merge_cells('B2:G2')
    ws['B2'] = CAT_DESC[cat]
    ws['B2'].font = Font(size=15, bold=True)
    ws.row_dimensions[2].height = 28

    for i, h in enumerate(HEADERS):
        c = ws.cell(row=4, column=i + 2, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 22

    r = 5
    for o in [x for x in data['pool'] if x['cat'] == cat]:
        vals = [o['id'], o['name'], o['desc'], o['tiers'], o['flavor'], '']
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=i + 2, value=v)
            c.font = thin
            c.alignment = Alignment(vertical='center', horizontal='center' if i == 0 else 'left')
            if r % 2 == 0:
                c.fill = PatternFill(start_color=GREY_BG, end_color=GREY_BG, fill_type='solid')
        ws.row_dimensions[r].height = 20
        r += 1
    ws.freeze_panes = 'B5'
    print(title, 'rows:', r - 5)

for cat in ['A', 'B', 'C', 'D']:
    fill_cat_sheet(cat)

# ─────────── Sheet 6: 段位之力 ───────────
ws = wb.create_sheet('段位之力')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
for col, w in zip('BCDE', [20, 30, 34, 30]):
    ws.column_dimensions[col].width = w
ws.merge_cells('B2:E2')
ws['B2'] = '段位之力 · 各段位专属增益（段位卡已下架，效果池保留备用）'
ws['B2'].font = Font(size=15, bold=True)
ws.row_dimensions[2].height = 28
for i, h in enumerate(['段位', '名称', '效果说明', '调整意见']):
    c = ws.cell(row=4, column=i + 2, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
r = 5
for o in data['tierfx']:
    for i, v in enumerate([o['tier'], o['name'], o['desc'], '']):
        c = ws.cell(row=r, column=i + 2, value=v)
        c.font = thin
        c.alignment = Alignment(vertical='center')
        if r % 2 == 0:
            c.fill = PatternFill(start_color=GREY_BG, end_color=GREY_BG, fill_type='solid')
    r += 1
ws.freeze_panes = 'B5'
print('段位之力 rows:', r - 5)

wb.save('人生选择配置清单.xlsx')
print('saved 人生选择配置清单.xlsx')
